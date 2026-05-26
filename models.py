"""
Classes de données pour la Coupe du Monde FIFA 2026.

Quatre dataclasses :
- Stadium : un stade hôte avec ses caractéristiques météo/géo.
- Camp    : un camp de base d'entraînement.
- Game    : un match (avec stade, équipes, date et fuseau).
- Team    : une équipe nationale, avec sa fédération et son camp.

Le module expose aussi `load_dataset()` qui lit `data_fifa.xlsx`
et renvoie un objet `Dataset` regroupant toutes les entités,
prêtes à être utilisées par `stats.py`.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl


# ---------------------------------------------------------------------------
# Dataclasses
# ---------------------------------------------------------------------------


@dataclass
class Stadium:
    city: str
    name: str
    country: str
    lat: float
    long: float
    altitude_m: float
    capacity: int
    temperature_c_june: float
    humidity_pct_june: float
    time_zone: str
    utc_offset: int  # offset en heures par rapport à UTC (en juin)


@dataclass
class Camp:
    site: str
    city: str
    country: str
    hotel: str
    lat: float
    long: float
    utc_offset: Optional[int] = None  # rempli par load_dataset à partir du lat/lon


@dataclass
class Game:
    match_id: int
    group: str
    date: date
    local_time: str
    team_1: str
    team_2: str
    stadium_name: str
    city: str
    country: str
    day: int
    utc_offset: int


@dataclass
class Team:
    name: str
    group: str
    rank_fifa: int
    points_fifa: float
    federation: str
    capital_lat: float
    capital_long: float
    utc_capital: int
    continent: str
    camp: Camp


# ---------------------------------------------------------------------------
# Conteneur principal
# ---------------------------------------------------------------------------


@dataclass
class Dataset:
    stadiums: Dict[str, Stadium]          # clé : nom du stade
    camps: List[Camp]                     # tous les camps disponibles
    games: List[Game]
    teams: Dict[str, Team]                # clé : nom de l'équipe

    def games_of(self, team_name: str) -> List[Game]:
        """Retourne tous les matchs de cette équipe, triés par date."""
        return sorted(
            [g for g in self.games if team_name in (g.team_1, g.team_2)],
            key=lambda g: g.date,
        )


# ---------------------------------------------------------------------------
# Chargement
# ---------------------------------------------------------------------------


def _to_float(value) -> float:
    if value is None or value == "":
        return float("nan")
    return float(value)


def _parse_date(value) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return datetime.strptime(str(value), "%Y-%m-%d").date()


def _utc_offset_from_latlon(lat: float, lon: float, on_date: date) -> int:
    """
    Calcule l'offset UTC (en heures, entier) à une date donnée,
    en utilisant timezonefinder + zoneinfo. Gère le DST.
    """
    from timezonefinder import TimezoneFinder
    from zoneinfo import ZoneInfo

    tf = TimezoneFinder()
    tz_name = tf.timezone_at(lat=lat, lng=lon)
    if tz_name is None:
        # fallback grossier : 1h par tranche de 15° de longitude
        return round(lon / 15)
    tz = ZoneInfo(tz_name)
    dt = datetime.combine(on_date, datetime.min.time()).replace(tzinfo=tz)
    return int(dt.utcoffset().total_seconds() // 3600)


def load_dataset(xlsx_path: str | Path) -> Dataset:
    """Lit le fichier `data_fifa.xlsx` et reconstruit toutes les entités."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # ---- Stadiums --------------------------------------------------------
    stadiums: Dict[str, Stadium] = {}
    rows = list(wb["stadiums"].iter_rows(values_only=True))
    header = rows[0]
    for row in rows[1:]:
        if row[0] is None:
            continue
        rec = dict(zip(header, row))
        s = Stadium(
            city=rec["city"],
            name=rec["stadium"],
            country=rec["country"],
            lat=_to_float(rec["let"]),      # le fichier source utilise "let"
            long=_to_float(rec["long"]),
            altitude_m=_to_float(rec["alt_m"]),
            capacity=int(rec["capacity"]),
            temperature_c_june=_to_float(rec["temp_C_june"]),
            humidity_pct_june=_to_float(rec["humidity_%_june"]),
            time_zone=rec["time_zone"],
            utc_offset=int(rec["UTC_Offset"]),
        )
        stadiums[s.name] = s

    # ---- Camps -----------------------------------------------------------
    camps: List[Camp] = []
    rows = list(wb["camps"].iter_rows(values_only=True))
    header = rows[0]
    # date de référence pour calculer le DST : 15 juin 2026 (milieu du tournoi)
    ref_date = date(2026, 6, 15)
    for row in rows[1:]:
        if row[0] is None:
            continue
        rec = dict(zip(header, row))
        lat, lon = _to_float(rec["latitude"]), _to_float(rec["longitude"])
        camp = Camp(
            site=rec["site_entrainement"],
            city=rec["ville"],
            country=rec["pays"],
            hotel=rec["hotel_associe"],
            lat=lat,
            long=lon,
            utc_offset=_utc_offset_from_latlon(lat, lon, ref_date),
        )
        camps.append(camp)

    # ---- Teams -----------------------------------------------------------
    teams: Dict[str, Team] = {}
    rows = list(wb["teams"].iter_rows(values_only=True))
    header = rows[0]
    for row in rows[1:]:
        if row[0] is None:
            continue
        rec = dict(zip(header, row))
        camp_lat = _to_float(rec["camp_lat"])
        camp_lon = _to_float(rec["camp_long"])
        camp = Camp(
            site=rec["camp_name"],
            city=rec["city_camp"],
            country=rec["country_camp"],
            hotel="",
            lat=camp_lat,
            long=camp_lon,
            utc_offset=_utc_offset_from_latlon(camp_lat, camp_lon, ref_date),
        )
        t = Team(
            name=rec["team"],
            group=rec["group"],
            rank_fifa=int(rec["rank_fifa"]),
            points_fifa=_to_float(rec["points_fifa"]),
            federation=rec["federation"],
            capital_lat=_to_float(rec["capital_lat"]),
            capital_long=_to_float(rec["capital_long"]),
            utc_capital=int(rec["UTC_capital"]),
            continent=rec["continent"],
            camp=camp,
        )
        teams[t.name] = t

    # ---- Games -----------------------------------------------------------
    games: List[Game] = []
    rows = list(wb["games"].iter_rows(values_only=True))
    header = rows[0]
    for row in rows[1:]:
        if row[0] is None:
            continue
        rec = dict(zip(header, row))
        g = Game(
            match_id=int(rec["#match"]),
            group=rec["group"],
            date=_parse_date(rec["date"]),
            local_time=str(rec["local_time"]),
            team_1=rec["team_1"],
            team_2=rec["team_2"],
            stadium_name=rec["stadium"],
            city=rec["city"],
            country=rec["country"],
            day=int(rec["day"]),
            utc_offset=int(rec["UTC_Offset"]),
        )
        games.append(g)

    return Dataset(stadiums=stadiums, camps=camps, games=games, teams=teams)


if __name__ == "__main__":
    ds = load_dataset(Path(__file__).parent / "data_fifa.xlsx")
    print(f"{len(ds.stadiums)} stades, {len(ds.camps)} camps, "
          f"{len(ds.teams)} équipes, {len(ds.games)} matchs")
